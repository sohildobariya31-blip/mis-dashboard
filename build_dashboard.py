"""
Purchase MIS Dashboard - Data Builder
=====================================
Reads the BOM sheet from the purchase record xlsx file and produces data.js
which is consumed by dashboard.html.

USAGE
-----
1. Update the Excel file (keep the same column layout in the BOM sheet).
2. Run:  python build_dashboard.py
3. Open dashboard.html in any browser.

You can change INPUT_FILE / SHEET_NAME below or pass them as CLI args:
    python build_dashboard.py "path\\to\\file.xlsx" BOM
"""

from __future__ import annotations
import json
import os
import sys
from datetime import datetime, date, time, timedelta

import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(HERE, "26-27-Purchase Record.xlsx")
SHEET_NAME = "BOM"
OUTPUT_JS = os.path.join(HERE, "data.js")

# Column header -> json key mapping (must match the BOM sheet header row)
COLUMN_MAP = {
    "Category (BOM)":      "category",
    "PO no":               "poNo",
    "PO Date":             "poDate",
    "Name of Supplier":    "supplier",
    "Module type":         "module",
    "Item Code":           "itemCode",
    "Material Description":"description",
    "Variant":             "variant",
    "Qty":                 "qty",
    "UOM":                 "uom",
    "Invoice no":          "invoiceNo",
    "Invoice/BL Date":     "invoiceDate",
    "BL no":               "blNo",
    "Status":              "status",
    "Project code":        "project",
    "COO":                 "coo",
    "Payment term":        "paymentTerm",
    "Shipping term":       "shippingTerm",
    "currency ":           "currency",
    "Basic Rate":          "basicRate",
    "Amount":              "amount",
    "Freight":             "freight",
    "GST%":                "gstPct",
    "Total Amount":        "totalAmount",
    "Payable amount":      "payable",
    "Debit Amount":        "debit",
    "Credit days":         "creditDays",
    "Due date":            "dueDate",
    "Payment status":      "paymentStatus",
    "Paid Amount":         "paid",
    "Outstanding":         "outstanding",
    "BOE no":              "boeNo",
    "BOE date":            "boeDate",
    "Landed cost/UOM":     "landedCost",
    "LC no":               "lcNo",
    "WR no":               "wrNo",
    "WR date":             "wrDate",
    "Lot no":              "lotNo",
    "QC status":           "qcStatus",
    "Remarks":             "remarks",
}

DATE_FIELDS = {"poDate", "invoiceDate", "dueDate", "boeDate", "wrDate"}


def cell_value(v):
    """Convert openpyxl cell value to JSON-serialisable form."""
    if v is None:
        return None
    if isinstance(v, datetime):
        # Treat the bogus 1900-02-14 placeholder as missing
        if v.year == 1900:
            return None
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        if v.year == 1900:
            return None
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, timedelta):
        return str(v)
    if isinstance(v, str):
        return v.strip() or None
    return v


def main():
    input_file = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else SHEET_NAME

    print(f"Reading: {input_file}  [sheet: {sheet_name}]")
    wb = openpyxl.load_workbook(input_file, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Read header row
    headers_raw = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = []
    for h in headers_raw:
        if h is None:
            headers.append(None)
            continue
        headers.append(str(h))

    # Validate required headers
    missing = [h for h in COLUMN_MAP if h not in headers]
    if missing:
        print("WARNING - these expected columns are missing:", missing)

    # Build column index -> json key
    col_to_key = {}
    for idx, h in enumerate(headers, start=1):
        if h in COLUMN_MAP:
            col_to_key[idx] = COLUMN_MAP[h]

    rows = []
    for r in range(2, ws.max_row + 1):
        record = {}
        non_empty = False
        for col_idx, key in col_to_key.items():
            v = cell_value(ws.cell(row=r, column=col_idx).value)
            if v is not None and v != "":
                non_empty = True
            record[key] = v
        if not non_empty:
            continue
        # Skip rows missing both PO no and Item Code (likely empty filler rows)
        if record.get("poNo") in (None, "") and not record.get("itemCode"):
            continue
        rows.append(record)

    # Coerce numeric fields just in case (Excel sometimes stores as text)
    numeric_fields = ["qty", "basicRate", "amount", "freight", "gstPct",
                      "totalAmount", "payable", "debit", "creditDays",
                      "paid", "outstanding", "landedCost"]
    for rec in rows:
        for f in numeric_fields:
            v = rec.get(f)
            if isinstance(v, str):
                try:
                    rec[f] = float(v.replace(",", ""))
                except Exception:
                    pass

    meta = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sourceFile": os.path.basename(input_file),
        "sheet": sheet_name,
        "rowCount": len(rows),
    }

    # Write data.js
    payload = {"meta": meta, "rows": rows}
    js = (
        "// Auto-generated by build_dashboard.py - do not edit by hand.\n"
        "window.PURCHASE_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n"
    )
    with open(OUTPUT_JS, "w", encoding="utf-8") as f:
        f.write(js)

    print(f"Wrote {OUTPUT_JS}  ({len(rows)} rows)")
    print("Open dashboard.html in a browser to view.")


if __name__ == "__main__":
    main()
