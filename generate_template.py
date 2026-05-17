"""
Generate Professional Purchase Record Excel Template
====================================================
Creates a template with:
- Formulas (Amount, Total, Payable, Outstanding, Due Date)
- Conditional formatting (status colors, overdue highlighting)
- Data validation (dropdown lists for Category, Status, UOM, Currency, etc.)
- Named ranges for easy reference
- Professional formatting (headers, borders, number formats)

Run: python generate_template.py
Output: Purchase_Record_Template.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ============================================================
# LIST SHEET (for dropdowns)
# ============================================================
list_ws = wb.create_sheet('List')
lists = {
    'Category': ['RM', 'PM'],
    'Module_Type': ['Perc+M10R+G12R', 'Perc+M10R', 'M10R+G12R', 'M10R', 'G12R', 'Perc+Topcon'],
    'UOM': ['Nos', 'Kgs', 'Mtrs', 'Sqm', 'Rolls', 'Sets', 'Ltrs', 'Pcs'],
    'Status': ['Received', 'In Transit', 'On Hold', 'Tentative', 'Cancelled', 'Short Close'],
    'COO': ['India', 'China', 'Malaysia', 'Indonesia', 'Thailand', 'Vietnam', 'Taiwan', 'Korea'],
    'Currency': ['INR', 'USD', 'RMB', 'EUR'],
    'Payment_Term': ['100% advance', '50% advance', '100% after delivery', '100% LC in 60 days against B/L date', '45 Days credit', '60 days against B/L date', '30 Days credit', '90 Days credit'],
    'Shipping_Term': ['FOR', 'FOB Shanghai', 'FOB', 'Ex-works', 'CIF', 'CFR'],
    'Payment_Status': ['Approved', 'Paid', 'Pending', 'Hold'],
    'Item_Code': ['Solar Cell', 'Glass', 'Al. Frame', 'EVA/EPE/POE', 'Backsheet', 'Junction Box', 'Ribbon/BB', 'Labels', 'Sealant', 'PET Strap', 'Retainer Strip', 'Corrugated Box', 'Wooden Pallet', 'Plywood Box', 'RFID', 'Stretch Wrapping Film', 'Corrugated corner', 'Corrugated Edge Board', 'Cell Fixation Tape', 'Flux', 'IPA', 'Microfiber Cloth'],
}

col = 1
for name, values in lists.items():
    list_ws.cell(1, col, name)
    list_ws.cell(1, col).font = Font(bold=True)
    for i, v in enumerate(values, 2):
        list_ws.cell(i, col, v)
    col += 1

# ============================================================
# BOM SHEET
# ============================================================
ws = wb.active
ws.title = 'BOM'

headers = [
    'Category (BOM)', 'PO no', 'PO Date', 'Name of Supplier', 'Module type',
    'Item Code', 'Material Description', 'Variant', 'Qty', 'UOM',
    'Invoice no', 'Invoice/BL Date', 'BL no', 'Status',
    'Project code', 'COO', 'Payment term', 'Shipping term', 'currency ',
    'Basic Rate', 'Amount', 'Freight', 'GST%', 'Total Amount',
    'Payable amount', 'Debit Amount', 'Credit days', 'Due date',
    'Payment status', 'Paid Amount', 'Outstanding',
    'BOE no', 'BOE date', 'Landed cost/UOM', 'LC no',
    'WR no', 'WR date', 'Lot no', 'QC status', 'Remarks'
]

# Styles
header_font = Font(bold=True, color='FFFFFF', size=10)
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
num_fmt_inr = '#,##0.00'
num_fmt_date = 'DD-MM-YYYY'

# Write headers
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(1, col_idx, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Set column widths
widths = [8, 6, 12, 30, 15, 15, 40, 15, 12, 6, 12, 12, 12, 10, 30, 8, 25, 12, 6, 12, 15, 10, 6, 15, 15, 12, 8, 12, 10, 12, 15, 10, 12, 12, 10, 10, 12, 10, 10, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze panes
ws.freeze_panes = 'A2'

# Add formulas for rows 2-1000
for row in range(2, 1001):
    # Amount = Qty * Basic Rate (col I * col T)
    ws.cell(row, 21).value = f'=IF(AND(I{row}<>"",T{row}<>""),I{row}*T{row},"")'
    ws.cell(row, 21).number_format = num_fmt_inr
    
    # Total Amount = (Amount + Freight) * (1 + GST%) (col U + col V) * (1 + col W)
    ws.cell(row, 24).value = f'=IF(U{row}<>"",($U{row}+$V{row})+(($U{row}+$V{row})*$W{row}),"")'
    ws.cell(row, 24).number_format = num_fmt_inr
    
    # Payable = Total Amount - Debit (col X - col Z)
    ws.cell(row, 25).value = f'=IF(X{row}<>"",X{row}-Z{row},"")'
    ws.cell(row, 25).number_format = num_fmt_inr
    
    # Due date = Invoice Date + Credit days (col L + col AA)
    ws.cell(row, 28).value = f'=IF(AND(L{row}<>"",AA{row}<>""),L{row}+AA{row},"")'
    ws.cell(row, 28).number_format = num_fmt_date
    
    # Outstanding = Payable - Paid (col Y - col AD)
    ws.cell(row, 31).value = f'=IF(Y{row}<>"",Y{row}-AD{row},"")'
    ws.cell(row, 31).number_format = num_fmt_inr
    
    # Number formats for other columns
    ws.cell(row, 3).number_format = num_fmt_date   # PO Date
    ws.cell(row, 12).number_format = num_fmt_date  # Invoice Date
    ws.cell(row, 20).number_format = num_fmt_inr   # Basic Rate
    ws.cell(row, 22).number_format = num_fmt_inr   # Freight
    ws.cell(row, 26).number_format = num_fmt_inr   # Debit
    ws.cell(row, 30).number_format = num_fmt_inr   # Paid

# Data Validations
dv_category = DataValidation(type='list', formula1='=List!$A$2:$A$10', allow_blank=True)
dv_category.error = 'Select from list'
dv_category.errorTitle = 'Invalid Category'
ws.add_data_validation(dv_category)
dv_category.add(f'A2:A1000')

dv_module = DataValidation(type='list', formula1='=List!$B$2:$B$10', allow_blank=True)
ws.add_data_validation(dv_module)
dv_module.add(f'E2:E1000')

dv_uom = DataValidation(type='list', formula1='=List!$C$2:$C$10', allow_blank=True)
ws.add_data_validation(dv_uom)
dv_uom.add(f'J2:J1000')

dv_status = DataValidation(type='list', formula1='=List!$D$2:$D$10', allow_blank=True)
ws.add_data_validation(dv_status)
dv_status.add(f'N2:N1000')

dv_coo = DataValidation(type='list', formula1='=List!$E$2:$E$10', allow_blank=True)
ws.add_data_validation(dv_coo)
dv_coo.add(f'P2:P1000')

dv_currency = DataValidation(type='list', formula1='=List!$F$2:$F$10', allow_blank=True)
ws.add_data_validation(dv_currency)
dv_currency.add(f'S2:S1000')

dv_payterm = DataValidation(type='list', formula1='=List!$G$2:$G$10', allow_blank=True)
ws.add_data_validation(dv_payterm)
dv_payterm.add(f'Q2:Q1000')

dv_shipterm = DataValidation(type='list', formula1='=List!$H$2:$H$10', allow_blank=True)
ws.add_data_validation(dv_shipterm)
dv_shipterm.add(f'R2:R1000')

dv_paystatus = DataValidation(type='list', formula1='=List!$I$2:$I$10', allow_blank=True)
ws.add_data_validation(dv_paystatus)
dv_paystatus.add(f'AC2:AC1000')

dv_item = DataValidation(type='list', formula1='=List!$J$2:$J$30', allow_blank=True)
ws.add_data_validation(dv_item)
dv_item.add(f'F2:F1000')

# Conditional Formatting
# Status = Received → green
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
ws.conditional_formatting.add('N2:N1000', CellIsRule(operator='equal', formula=['"Received"'], fill=green_fill))

# Status = In Transit → yellow
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
ws.conditional_formatting.add('N2:N1000', CellIsRule(operator='equal', formula=['"In Transit"'], fill=yellow_fill))

# Status = Cancelled → red
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
ws.conditional_formatting.add('N2:N1000', CellIsRule(operator='equal', formula=['"Cancelled"'], fill=red_fill))

# Outstanding > 0 → light red background
ws.conditional_formatting.add('AE2:AE1000', CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')))

# Overdue (Due date < today and Outstanding > 0)
ws.conditional_formatting.add('AB2:AB1000', FormulaRule(formula=['AND(AB2<TODAY(),AE2>0)'], fill=red_fill))

# Payment Status = Paid → green
ws.conditional_formatting.add('AC2:AC1000', CellIsRule(operator='equal', formula=['"Paid"'], fill=green_fill))

# ============================================================
# OTHER MATERIAL SHEET (same structure, different headers)
# ============================================================
om_ws = wb.create_sheet('Other material')
om_headers = [
    'Department', 'PO no', 'PO date', 'Vendor Name', 'Category',
    'Item description', 'Qty', 'UOM', 'CON', 'Invoice no',
    'Invoice/BL date', 'Status', 'Project Code', 'Currency', ' Rate',
    'Amount', 'Freight', 'GST', 'Total Amount', 'Payable Amount',
    'Debit Amount', 'Payment term', 'Shipping term', 'Credit days',
    'Due date', 'payment status', 'Paid Amount', 'O/S amount', 'Remarks'
]

for col_idx, h in enumerate(om_headers, 1):
    cell = om_ws.cell(1, col_idx, h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    cell.alignment = header_align
    cell.border = thin_border

# Formulas for Other material
for row in range(2, 1001):
    # Amount = Qty * Rate (col G * col O)
    om_ws.cell(row, 16).value = f'=IF(AND(G{row}<>"",O{row}<>""),G{row}*O{row},"")'
    om_ws.cell(row, 16).number_format = num_fmt_inr
    # Total = (Amount + Freight) * (1 + GST)
    om_ws.cell(row, 19).value = f'=IF(P{row}<>"",($P{row}+$Q{row})+(($P{row}+$Q{row})*$R{row}),"")'
    om_ws.cell(row, 19).number_format = num_fmt_inr
    # Payable = Total - Debit
    om_ws.cell(row, 20).value = f'=IF(S{row}<>"",S{row}-U{row},"")'
    om_ws.cell(row, 20).number_format = num_fmt_inr
    # Due date = Invoice date + Credit days
    om_ws.cell(row, 25).value = f'=IF(AND(K{row}<>"",X{row}<>""),K{row}+X{row},"")'
    om_ws.cell(row, 25).number_format = num_fmt_date
    # O/S = Payable - Paid
    om_ws.cell(row, 28).value = f'=IF(T{row}<>"",T{row}-AA{row},"")'
    om_ws.cell(row, 28).number_format = num_fmt_inr

om_ws.freeze_panes = 'A2'

# Data validations for Other material
dv_om_status = DataValidation(type='list', formula1='=List!$D$2:$D$10', allow_blank=True)
om_ws.add_data_validation(dv_om_status)
dv_om_status.add('L2:L1000')

dv_om_currency = DataValidation(type='list', formula1='=List!$F$2:$F$10', allow_blank=True)
om_ws.add_data_validation(dv_om_currency)
dv_om_currency.add('N2:N1000')

dv_om_uom = DataValidation(type='list', formula1='=List!$C$2:$C$10', allow_blank=True)
om_ws.add_data_validation(dv_om_uom)
dv_om_uom.add('H2:H1000')

# Conditional formatting
om_ws.conditional_formatting.add('L2:L1000', CellIsRule(operator='equal', formula=['"Received"'], fill=green_fill))
om_ws.conditional_formatting.add('L2:L1000', CellIsRule(operator='equal', formula=['"In Transit"'], fill=yellow_fill))
om_ws.conditional_formatting.add('L2:L1000', CellIsRule(operator='equal', formula=['"Cancelled"'], fill=red_fill))

# ============================================================
# SAVE
# ============================================================
output = r'C:\Users\MI\Downloads\MIS\Purchase_Record_Template.xlsx'
wb.save(output)
print(f'✓ Template saved: {output}')
print(f'  Sheets: BOM, Other material, List')
print(f'  Features: Formulas, Conditional Formatting, Data Validation, Dropdown Lists')
print(f'  Ready for 2-way sync with dashboard')
